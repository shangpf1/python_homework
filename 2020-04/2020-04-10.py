# 数据驱动（xlsx数据驱动）

import requests
import pytest
import os
from openpyxl import  load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# 文件路径
dir_path = os.path.dirname(os.path.dirname(__file__))
execfile = os.path.join(dir_path, 'data/data.xlsx')

workbook=load_workbook(execfile)
ws:Worksheet=workbook['topic_data']
exceldata=[]
for row in ws.iter_rows(min_row=2,max_row=6,min_col=1,max_col=5,values_only=True):
    # print(row)
    exceldata.append(row)



"""
 [['0418178a-b80c-4e15-aa8f-bab03a3491cb', '1111111111', 'ask', '22222222222', '错误的accessToken'], ['49b2e830-4305-475d-b6b5-52287cc5daaa', '', 'ask', '2222222222', '标题不能为空'], ['49b2e830-4305-475d-b6b5-52287cc5daaa', '1', 'ask', '2222222222', '标题字数太多或太少'], ['49b2e830-4305-475d-b6b5-52287cc5daaa', '1111111111', '', '2222222222', '必须选择一个版块'], ['49b2e830-4305-475d-b6b5-52287cc5daaa', '1111111111', 'ask', '', '内容不可为空']]
  ||   数据格式转换一下
  \/
[{"token":"0418178a-b80c-4e15-aa8f-bab03a3491cb","tab":"ask"},{....}]
"""

@pytest.fixture(params=exceldata)
def data(request):
    return request.param

base_url="http://39.107.96.138:3000/api/v1/"

def test_topics(data):
    url = base_url + 'topics'
    testdata={
        "accesstoken": data[0],
        "title": data[1],
        "tab": data[2],
        "content": data[3]
    }

    r = requests.post(url,json=testdata)
    jsonData = r.json()
    assert jsonData['error_msg'] == data[4]

